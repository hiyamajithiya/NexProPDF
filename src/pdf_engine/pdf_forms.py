"""
PDF Forms - Create, fill, flatten, and export form data
"""

import fitz  # PyMuPDF
import pikepdf
from typing import List, Dict, Optional, Tuple
from pathlib import Path
from src.utilities.logger import get_logger


class PDFForms:
    """PDF forms operations"""

    def __init__(self):
        self.logger = get_logger()

    def create_text_field(self, pdf_document, page_num: int, field_name: str,
                         rect: Tuple[float, float, float, float],
                         default_value: str = "", font_size: int = 12) -> bool:
        """
        Create text field in PDF

        Args:
            pdf_document: PyMuPDF document
            page_num: Page number
            field_name: Field name
            rect: Field rectangle (x0, y0, x1, y1)
            default_value: Default text value
            font_size: Font size

        Returns:
            True if successful
        """
        try:
            page = pdf_document[page_num]

            # Create text widget
            widget = fitz.Widget()
            widget.field_name = field_name
            widget.field_type = fitz.PDF_WIDGET_TYPE_TEXT
            widget.rect = fitz.Rect(rect)
            widget.field_value = default_value
            widget.text_fontsize = font_size
            widget.text_color = (0, 0, 0)  # Black
            widget.fill_color = (1, 1, 1)  # White background

            # Add widget to page
            page.add_widget(widget)

            self.logger.info(f"Created text field '{field_name}' on page {page_num}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating text field: {e}")
            return False

    def create_checkbox(self, pdf_document, page_num: int, field_name: str,
                       rect: Tuple[float, float, float, float],
                       checked: bool = False) -> bool:
        """
        Create checkbox in PDF

        Args:
            pdf_document: PyMuPDF document
            page_num: Page number
            field_name: Field name
            rect: Field rectangle
            checked: Initial checked state

        Returns:
            True if successful
        """
        try:
            page = pdf_document[page_num]

            # Create checkbox widget
            widget = fitz.Widget()
            widget.field_name = field_name
            widget.field_type = fitz.PDF_WIDGET_TYPE_CHECKBOX
            widget.rect = fitz.Rect(rect)
            widget.field_value = checked

            # Add widget to page
            page.add_widget(widget)

            self.logger.info(f"Created checkbox '{field_name}' on page {page_num}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating checkbox: {e}")
            return False

    def create_radio_button(self, pdf_document, page_num: int, field_name: str,
                           rect: Tuple[float, float, float, float],
                           button_value: str) -> bool:
        """
        Create radio button in PDF

        Args:
            pdf_document: PyMuPDF document
            page_num: Page number
            field_name: Field name (same for all buttons in group)
            rect: Field rectangle
            button_value: Value when selected

        Returns:
            True if successful
        """
        try:
            page = pdf_document[page_num]

            # Create radio button widget
            widget = fitz.Widget()
            widget.field_name = field_name
            widget.field_type = fitz.PDF_WIDGET_TYPE_RADIOBUTTON
            widget.rect = fitz.Rect(rect)
            widget.button_caption = button_value

            # Add widget to page
            page.add_widget(widget)

            self.logger.info(f"Created radio button '{field_name}' on page {page_num}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating radio button: {e}")
            return False

    def create_dropdown(self, pdf_document, page_num: int, field_name: str,
                       rect: Tuple[float, float, float, float],
                       options: List[str], default_index: int = 0) -> bool:
        """
        Create dropdown field in PDF

        Args:
            pdf_document: PyMuPDF document
            page_num: Page number
            field_name: Field name
            rect: Field rectangle
            options: List of dropdown options
            default_index: Default selected index

        Returns:
            True if successful
        """
        try:
            page = pdf_document[page_num]

            # Create combo box widget
            widget = fitz.Widget()
            widget.field_name = field_name
            widget.field_type = fitz.PDF_WIDGET_TYPE_COMBOBOX
            widget.rect = fitz.Rect(rect)
            widget.choice_values = options
            if 0 <= default_index < len(options):
                widget.field_value = options[default_index]

            # Add widget to page
            page.add_widget(widget)

            self.logger.info(f"Created dropdown '{field_name}' on page {page_num}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating dropdown: {e}")
            return False

    def get_form_fields(self, pdf_document) -> List[Dict]:
        """
        Get all form fields from PDF

        Args:
            pdf_document: PyMuPDF document

        Returns:
            List of field dictionaries
        """
        fields = []

        try:
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]

                # Get widgets (form fields)
                for widget in page.widgets():
                    field_info = {
                        'page': page_num,
                        'name': widget.field_name,
                        'type': widget.field_type_string,
                        'value': widget.field_value,
                        'rect': tuple(widget.rect)
                    }
                    fields.append(field_info)

            self.logger.info(f"Found {len(fields)} form fields")
            return fields

        except Exception as e:
            self.logger.error(f"Error getting form fields: {e}")
            return []

    def fill_form_field(self, pdf_document, field_name: str, value) -> bool:
        """
        Fill form field with value

        Args:
            pdf_document: PyMuPDF document
            field_name: Field name to fill
            value: Value to set

        Returns:
            True if successful
        """
        try:
            filled = False

            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]

                for widget in page.widgets():
                    if widget.field_name == field_name:
                        widget.field_value = value
                        widget.update()
                        filled = True

            if filled:
                self.logger.info(f"Filled field '{field_name}' with value: {value}")
            else:
                self.logger.warning(f"Field '{field_name}' not found")

            return filled

        except Exception as e:
            self.logger.error(f"Error filling form field: {e}")
            return False

    def flatten_form(self, input_file: str, output_file: str) -> bool:
        """
        Flatten PDF form (make fields non-editable)

        Args:
            input_file: Input PDF path
            output_file: Output PDF path

        Returns:
            True if successful
        """
        try:
            with pikepdf.open(input_file) as pdf:
                # Flatten form fields
                if '/AcroForm' in pdf.Root:
                    # Remove interactive form
                    del pdf.Root['/AcroForm']

                pdf.save(output_file)

            self.logger.info(f"Flattened form: {output_file}")
            return True

        except Exception as e:
            self.logger.error(f"Error flattening form: {e}")
            return False

    def export_form_data(self, pdf_document, output_format: str = "dict") -> Optional[Dict]:
        """
        Export form data to dictionary or CSV

        Args:
            pdf_document: PyMuPDF document
            output_format: Format ('dict' or 'csv')

        Returns:
            Dictionary of form data or None
        """
        try:
            form_data = {}

            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]

                for widget in page.widgets():
                    field_name = widget.field_name
                    field_value = widget.field_value

                    form_data[field_name] = field_value

            self.logger.info(f"Exported {len(form_data)} form fields")
            return form_data

        except Exception as e:
            self.logger.error(f"Error exporting form data: {e}")
            return None

    def export_to_excel(self, pdf_document, output_file: str) -> bool:
        """
        Export form data to Excel file

        Args:
            pdf_document: PyMuPDF document
            output_file: Output Excel file path

        Returns:
            True if successful
        """
        try:
            import openpyxl
            from openpyxl import Workbook

            # Get form data
            form_data = self.export_form_data(pdf_document)

            if not form_data:
                return False

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Form Data"

            # Add headers
            ws['A1'] = "Field Name"
            ws['B1'] = "Field Value"

            # Add data
            row = 2
            for field_name, field_value in form_data.items():
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = str(field_value)
                row += 1

            # Save workbook
            wb.save(output_file)

            self.logger.info(f"Exported form data to Excel: {output_file}")
            return True

        except ImportError:
            self.logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            return False

    def import_form_data(self, pdf_document, data: Dict) -> bool:
        """
        Import form data from dictionary

        Args:
            pdf_document: PyMuPDF document
            data: Dictionary of field_name: value pairs

        Returns:
            True if successful
        """
        try:
            filled_count = 0

            for field_name, value in data.items():
                if self.fill_form_field(pdf_document, field_name, value):
                    filled_count += 1

            self.logger.info(f"Imported {filled_count}/{len(data)} form fields")
            return filled_count > 0

        except Exception as e:
            self.logger.error(f"Error importing form data: {e}")
            return False
